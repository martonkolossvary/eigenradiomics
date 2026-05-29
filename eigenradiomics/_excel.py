"""Shared helper for writing polished, formatted multi-sheet Excel workbooks."""

from __future__ import annotations

from collections.abc import Callable, Mapping
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# A callable mapping (column_name, cell_value) -> openpyxl number format string
# (or None to leave the cell's format unchanged).  Only invoked for finite
# numeric cells.
NumberFormat = Callable[[str, float], "str | None"]


def write_styled_workbook(
    sheets: Mapping[str, pd.DataFrame],
    path: str | Path,
    number_format: NumberFormat,
) -> None:
    """Write *sheets* to a styled Excel workbook at *path*.

    Each sheet gets a dark-navy header row with frozen panes, auto-filters,
    thin borders, alignment by column type, and auto-fit column widths.  The
    per-cell numeric display format is delegated to *number_format* so callers
    can apply domain-specific rules.

    Parameters
    ----------
    sheets : mapping of str to DataFrame
        Ordered mapping of sheet name to table.
    path : str or Path
        Target ``.xlsx`` path.
    number_format : callable
        ``(column_name, value) -> format_string | None`` applied to finite
        numeric cells.
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    navy_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    white_bold = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)

    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    thin_border_side = Side(style="thin", color="D3D3D3")
    data_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side,
    )

    for sheet_name, df in sheets.items():
        ws = wb.create_sheet(title=sheet_name)

        headers = list(df.columns)
        ws.append(headers)

        # Format header row
        ws.row_dimensions[1].height = 28
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = navy_fill
            cell.font = white_bold
            cell.alignment = header_align

        # Write data rows
        for _, row in df.iterrows():
            ws.append(list(row))

        # Format data cells
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = regular_font
                cell.border = data_border

                col_name = headers[col_idx - 1]
                val = cell.value

                if isinstance(val, bool):
                    cell.alignment = center_align
                elif col_name == "feature":
                    cell.alignment = left_align
                    cell.font = bold_font
                elif isinstance(val, (int, float)) and not (
                    isinstance(val, float) and np.isnan(val)
                ):
                    cell.alignment = right_align
                    fmt = number_format(col_name, float(val))
                    if fmt is not None:
                        cell.number_format = fmt
                else:
                    cell.alignment = center_align

        # Freeze top row
        ws.freeze_panes = "A2"

        # Enable auto-filters
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

        # Auto-fit column widths with padding
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = cell.value
                if val is not None:
                    s = f"{val:.4f}" if isinstance(val, float) else str(val)
                    max_len = max(max_len, len(s))
            ws.column_dimensions[col_letter].width = max(max_len + 4, 11)

    wb.save(path)
