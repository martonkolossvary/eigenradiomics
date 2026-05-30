"""Unit tests for the batch effect diagnostics framework."""

from __future__ import annotations

import tempfile
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import pytest

from eigenradiomics.batch_effects import (
    anova_effect,
    compute_batch_effects,
    kruskal_effect,
    levene_test,
    permanova_euclidean,
    plot_batch_effects,
    write_batch_effects_excel,
)
from eigenradiomics.preprocessing import RadiomicsPrepTransformer

try:
    import inmoose  # noqa: F401

    _HAS_INMOOSE = True
except ImportError:
    _HAS_INMOOSE = False


def _batch_df(n_feat: int = 6, per_batch: int = 12, seed: int = 0, with_effect: bool = True):
    """Synthetic 3-center radiomics matrix with an optional batch shift."""
    rng = np.random.default_rng(seed)
    batches = np.repeat(["A", "B", "C"], per_batch)
    n = len(batches)
    X = rng.standard_normal((n, n_feat))
    if with_effect:
        shift = np.array([{"A": 0.0, "B": 1.5, "C": 3.0}[b] for b in batches])
        X[:, : n_feat // 2] += shift[:, None]
    cols = [f"original__f{i}" for i in range(n_feat)]
    df = pd.DataFrame(X, columns=cols, index=[f"S{i}" for i in range(n)])
    return df, pd.Series(batches, index=df.index)


def test_radiomics_prep_transformer():
    """Verify that RadiomicsPrepTransformer clips, transforms, scales, and carries NaNs."""
    df = pd.DataFrame(
        {
            "feat_1": [1.0, 2.0, np.nan, 4.0, 100.0],  # Outlier 100.0
            "feat_2": [10.0, 10.0, 10.0, 10.0, 10.0],  # Constant column
        }
    )

    transformer = RadiomicsPrepTransformer(
        winsor_lower=0.1, winsor_upper=0.9, skip_yeo_johnson=False
    )
    transformer.fit(df)

    # 1. Inspect fitted attributes
    assert transformer.n_features_in_ == 2
    assert len(transformer.winsor_bounds_) == 2

    # Check winsor bounds (1st feature winsorized at 10% and 90% -> roughly 1.3 and 70.8)
    low, high = transformer.winsor_bounds_[0]
    assert low < 2.0
    assert high < 100.0

    # Constant column power transformer should be None to avoid errors
    assert transformer.power_transformers_[1] is None

    # 2. Inspect transform output
    df_trans = transformer.transform(df)
    assert isinstance(df_trans, pd.DataFrame)
    assert np.isnan(df_trans.loc[2, "feat_1"])  # NaN must be preserved!
    assert df_trans.loc[4, "feat_1"] < 2.0  # Clipped/winsorized outlier should be small

    # Constant column should be returned as raw or zero-centered safely
    assert np.allclose(df_trans["feat_2"].dropna(), 10.0)


def test_batch_effects_anova_kruskal_levene():
    """Verify the feature-level ANOVA, Kruskal-Wallis, and Levene calculators."""
    g1 = np.array([1.0, 2.0, 3.0])
    g2 = np.array([4.0, 5.0, 6.0])

    f, p, eta2 = anova_effect([g1, g2])
    assert f > 0
    assert 0 < p < 1
    assert 0.5 < eta2 <= 1.0

    h, p_k, eps2 = kruskal_effect([g1, g2])
    assert h > 0
    assert 0 < p_k < 1
    assert eps2 > 0

    l_stat, p_l = levene_test([g1, g2])
    assert not np.isnan(l_stat)
    assert 0 < p_l <= 1.0


def test_permanova_pseudo_f():
    """Verify the PERMANOVA pseudo-F permutation calculation."""
    df = pd.DataFrame(
        {
            "PC1": [1.0, 1.1, 1.2, 5.0, 5.1, 5.2],
            "PC2": [2.0, 2.1, 2.2, 8.0, 8.1, 8.2],
        }
    )
    batch = pd.Series(["A", "A", "A", "B", "B", "B"])

    f, r2, p = permanova_euclidean(df, batch, permutations=99, random_state=42)
    assert f > 5.0
    assert r2 > 0.8
    assert 0 < p <= 0.1  # Significant separation


def test_compute_batch_effects_qc_and_alignment():
    """Test strict index QC alignment, reordering, and shape consistency validations."""
    X = pd.DataFrame(
        {"feat_1": [1, 2, 3], "feat_2": [4, 5, 6]},
        index=["S1", "S2", "S3"],
    )
    pd.Series(["A", "B", "A"], index=["S1", "S2", "S3"])

    # 1. Index mismatch should raise ValueError
    batch_bad = pd.Series(["A", "B", "A"], index=["S1", "S2", "S4"])
    with pytest.raises(ValueError, match="Indices of X and batch do not match"):
        compute_batch_effects(X, batch_bad)

    # 2. Scrambled index should automatically reorder and succeed
    batch_scrambled = pd.Series(["B", "A", "A"], index=["S2", "S1", "S3"])
    results = compute_batch_effects(
        X,
        batch_scrambled,
        permutations=10,
        min_valid_samples=3,
        min_valid_per_batch=1,
        min_batches_per_feature=2,
        no_combat=True,
    )
    assert "feature_stats" in results

    # 3. Shape mismatch on arrays
    arr = X.to_numpy()
    batch_arr_bad = np.array(["A", "B"])
    with pytest.raises(ValueError, match="Length of batch"):
        compute_batch_effects(arr, batch_arr_bad)


def test_compute_batch_effects_selectors_and_combat():
    """Test feature selector filtering, custom preprocessors, and ComBat sensitivity."""
    # Build standard wide radiomics features columns config__feature_key
    cols = [
        "original__Energy",
        "original__Entropy",
        "original__Autocorrelation",
        "wavelet__ClusterShade",
    ]
    df = pd.DataFrame(
        np.random.default_rng(42).standard_normal((30, 4)),
        columns=cols,
        index=[f"S{i}" for i in range(30)],
    )
    # 3 batches
    batch = pd.Series(["A", "B", "C"] * 10, index=df.index)

    # Pre-select feature catalog
    catalog = pd.DataFrame(
        {
            "feature_key": ["Energy", "Entropy", "Autocorrelation", "ClusterShade"],
            "config": ["original", "original", "original", "wavelet"],
            "family": ["firstorder", "firstorder", "glcm", "glcm"],
            "family_group": ["texture", "texture", "texture", "texture"],
        }
    )

    # Execute with selectors and combat sensitivity
    results = compute_batch_effects(
        df,
        batch,
        features=["*Energy", "*Entropy"],
        catalog=catalog,
        permutations=10,
        no_combat=False,
    )

    # Check results
    assert "dataset_summary" in results
    assert "batch_counts" in results
    assert "global_diagnostics" in results
    assert "feature_stats" in results

    # Only Energy and Entropy should be analyzed
    feature_stats = results["feature_stats"]
    assert len(feature_stats) == 2
    assert "original__Energy" in feature_stats["feature"].tolist()
    assert "original__Entropy" in feature_stats["feature"].tolist()

    # ComBat sensitivity sheets appear only when the optional inmoose extra is installed.
    if _HAS_INMOOSE:
        assert "combat_feature_stats" in results
        assert "combat_adjustment_notes" in results
    else:
        assert "combat_feature_stats" not in results


def test_excel_and_plot_batch_effects():
    """Test Excel writing and high-contrast accessibility figures."""
    cols = ["original__Energy", "original__Entropy"]
    df = pd.DataFrame(
        np.random.default_rng(42).standard_normal((30, 2)),
        columns=cols,
        index=[f"S{i}" for i in range(30)],
    )
    batch = pd.Series(["A", "B", "C"] * 10, index=df.index)

    results = compute_batch_effects(df, batch, permutations=10, no_combat=False)

    with tempfile.TemporaryDirectory() as tmp_dir:
        # 1. Excel Workbook export check
        xlsx_path = Path(tmp_dir) / "batch_report.xlsx"
        write_batch_effects_excel(results, xlsx_path)

        assert xlsx_path.exists()
        wb = openpyxl.load_workbook(xlsx_path)
        assert "dataset_summary" in wb.sheetnames
        assert "global_diagnostics" in wb.sheetnames
        assert "feature_stats" in wb.sheetnames

        ws = wb["feature_stats"]
        assert ws.cell(row=1, column=1).value == "feature"
        assert ws.cell(row=2, column=1).value == "original__Energy"

        # 2. Plotting accessibility visuals check
        fig_path = Path(tmp_dir) / "batch_effect_plot.png"
        fig = plot_batch_effects(results, path=fig_path, primary_alpha=0.05)

        assert fig_path.exists()
        assert isinstance(fig, plt.Figure)
        # The ComBat PCA panel is only added when inmoose is installed.
        assert len(fig.axes) == (3 if _HAS_INMOOSE else 2)
        plt.close(fig)


def test_batch_effects_skips_non_numeric_metadata_without_selectors():
    """Without selectors, non-numeric metadata columns must be skipped, not crash."""
    X = pd.DataFrame(
        {
            "PatientID": [f"p{i}" for i in range(6)],
            "original__Energy": [1.0, 2.0, 3.0, 4.0, 5.0, 6.0],
            "original__Entropy": [6.0, 5.0, 4.0, 3.0, 2.0, 1.0],
        }
    )
    batch = pd.Series(["A", "B"] * 3, index=X.index)

    results = compute_batch_effects(
        X,
        batch,
        permutations=5,
        no_combat=True,
        min_valid_samples=3,
        min_valid_per_batch=1,
        min_batches_per_feature=1,
    )
    analyzed = results["feature_stats"]["feature"].tolist()
    assert "PatientID" not in analyzed
    assert set(analyzed) == {"original__Energy", "original__Entropy"}


def test_array_input_runs():
    df, batch = _batch_df(with_effect=False)
    res = compute_batch_effects(df.to_numpy(), batch.to_numpy(), permutations=10, no_combat=True)
    assert "feature_stats" in res


def test_no_numeric_features_raises():
    X = pd.DataFrame({"label": list("abcdef")})
    batch = pd.Series(["A", "B", "C", "A", "B", "C"])
    with pytest.raises(ValueError, match="No features selected"):
        compute_batch_effects(X, batch)


def test_pipeline_returning_ndarray():
    from sklearn.preprocessing import StandardScaler

    df, batch = _batch_df(with_effect=False)
    res = compute_batch_effects(
        df, batch, pipeline=StandardScaler(), permutations=10, no_combat=True
    )
    assert "feature_stats" in res


def test_all_features_fail_qc_raises():
    df, batch = _batch_df()
    with pytest.raises(ValueError, match="No features passed QC"):
        compute_batch_effects(df, batch, no_combat=True, min_valid_samples=10_000)


def test_global_missing_complete_cases():
    df, batch = _batch_df()
    res = compute_batch_effects(
        df, batch, global_missing_strategy="complete-cases", permutations=10, no_combat=True
    )
    assert "global_diagnostics" in res


def test_global_missing_median_impute():
    df, batch = _batch_df()
    df.iloc[0, 0] = np.nan
    res = compute_batch_effects(
        df, batch, global_missing_strategy="median-impute", permutations=10, no_combat=True
    )
    assert "global_diagnostics" in res


def test_invalid_global_missing_strategy():
    df, batch = _batch_df()
    with pytest.raises(ValueError, match="Unsupported global missing strategy"):
        compute_batch_effects(df, batch, global_missing_strategy="bogus", no_combat=True)


def test_global_matrix_too_small_fallback():
    df, batch = _batch_df(n_feat=4)
    for j in range(4):
        df.iloc[j, j] = np.nan  # each feature gets a NaN -> complete-features drops all columns
    res = compute_batch_effects(
        df, batch, global_missing_strategy="complete-features", permutations=10, no_combat=True
    )
    assert int(res["global_diagnostics"]["pc_components_used"].iloc[0]) == 0


def _install_fake_inmoose(monkeypatch, norm_fn):
    """Inject a fake ``inmoose.pycombat`` so the ComBat path runs without the real library."""
    import sys
    import types

    fake = types.ModuleType("inmoose")
    fake_pc = types.ModuleType("inmoose.pycombat")
    fake_pc.pycombat_norm = norm_fn
    fake.pycombat = fake_pc
    monkeypatch.setitem(sys.modules, "inmoose", fake)
    monkeypatch.setitem(sys.modules, "inmoose.pycombat", fake_pc)


def test_combat_path_mocked_dataframe(monkeypatch):
    def norm(data, batch=None, covar_mod=None, par_prior=True, mean_only=False, ref_batch=None):
        return data  # identity correction (features x samples DataFrame)

    _install_fake_inmoose(monkeypatch, norm)
    df, batch = _batch_df()
    res = compute_batch_effects(df, batch, permutations=10, no_combat=False)
    assert "combat_feature_stats" in res
    assert "combat_adjustment_notes" in res

    # With ComBat results present, the plot gains the "PCA After ComBat" panel.
    fig = plot_batch_effects(res)
    assert len(fig.axes) == 3
    plt.close(fig)


def test_combat_path_mocked_ndarray_and_nonfinite(monkeypatch):
    def norm(data, batch=None, covar_mod=None, par_prior=True, mean_only=False, ref_batch=None):
        arr = np.asarray(data, dtype=float).copy()
        arr[0, 0] = np.nan  # exercise the non-finite replacement branch
        return arr  # ndarray (not DataFrame) -> exercises that branch

    _install_fake_inmoose(monkeypatch, norm)
    df, batch = _batch_df()
    covars = pd.DataFrame({"sex": [["M", "F"][i % 2] for i in range(len(df))]}, index=df.index)
    res = compute_batch_effects(
        df, batch, permutations=10, no_combat=False, combat_covariates=covars
    )
    assert "combat_feature_stats" in res
    assert int(res["combat_adjustment_notes"]["nonfinite_values_replaced"].iloc[0]) >= 1


def test_combat_import_failure(monkeypatch):
    import builtins

    real_import = builtins.__import__

    def block(name, *args, **kwargs):
        if name.startswith("inmoose"):
            raise ImportError("blocked for test")
        return real_import(name, *args, **kwargs)

    monkeypatch.setattr(builtins, "__import__", block)
    df, batch = _batch_df()
    with pytest.warns(UserWarning, match="ComBat sensitivity diagnostics skipped"):
        res = compute_batch_effects(df, batch, permutations=10, no_combat=False)
    assert "combat_feature_stats" not in res
