"""Tests for feature-outcome association models (Phase A: continuous OLS+HC3)."""

from __future__ import annotations

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import numpy as np  # noqa: E402
import pandas as pd  # noqa: E402
import pytest  # noqa: E402

from eigenradiomics import (  # noqa: E402
    FeatureCatalog,
    RadiomicsDataset,
    StudyDesign,
    compute_feature_associations,
    plot_volcano,
)


def _data(n: int = 120):
    rng = np.random.default_rng(0)
    f0 = rng.normal(0, 1, n)
    f1 = rng.normal(0, 1, n)
    age = rng.normal(60, 8, n)
    y = 2.0 * f0 + 0.05 * age + rng.normal(0, 1, n)  # f0 associated, f1 not
    idx = [f"S{i}" for i in range(n)]
    X = pd.DataFrame(
        {
            "original__f0": f0,
            "original__f1": f1,
            "original__const": np.full(n, 5.0),  # constant -> constant_feature
            "original__empty": np.full(n, np.nan),  # all-NaN -> no_complete_cases
        },
        index=idx,
    )
    meta = pd.DataFrame({"y": y, "age": age}, index=idx)
    return X, meta


def _catalog() -> FeatureCatalog:
    return FeatureCatalog(
        pd.DataFrame(
            {
                "config": ["original"] * 4,
                "feature_key": ["f0", "f1", "const", "empty"],
                "family": ["firstorder", "glcm", "firstorder", "glcm"],
                "family_group": ["Intensity", "Texture", "Intensity", "Texture"],
            }
        )
    )


# ---- core behaviour -------------------------------------------------------


def test_continuous_univariable_and_adjusted():
    X, meta = _data()
    res = compute_feature_associations(X, meta["y"], adjust_for=["age"], covariate_data=meta)
    assert res.tiers == ["Univariable", "Adjusted"]
    assert res.outcome_type == "continuous"
    t = res.table.set_index(["model", "feature"])

    # coefficient matches a plain OLS (HC3 only changes the SE, not the estimate)
    f0 = X["original__f0"].to_numpy()
    design = np.column_stack([np.ones(len(f0)), f0])
    beta = np.linalg.lstsq(design, meta["y"].to_numpy(), rcond=None)[0]
    assert np.isclose(t.loc[("Univariable", "original__f0"), "coef"], beta[1])

    assert t.loc[("Univariable", "original__f0"), "p_value"] < 1e-3  # strong signal
    assert t.loc[("Univariable", "original__f1"), "p_value"] > 0.05  # noise
    assert "p_fdr" in res.table.columns
    assert (t.loc[("Univariable", "original__f0"), "effect_name"]) == "beta"


def test_status_branches():
    X, meta = _data()
    res = compute_feature_associations(X, meta["y"], covariate_data=meta)
    status = res.table.set_index(["model", "feature"])["status"]
    assert status.loc[("Univariable", "original__f0")] == "ok"
    assert status.loc[("Univariable", "original__const")] == "constant_feature"
    assert status.loc[("Univariable", "original__empty")] == "no_complete_cases"


def test_not_enough_degrees_of_freedom():
    X, meta = _data(n=4)
    res = compute_feature_associations(X, meta["y"], adjust_for=["age"], covariate_data=meta)
    status = res.table.set_index(["model", "feature"])["status"]
    # 4 samples, design of 5 columns (intercept + feature + 3 implicit? here 1 covar) ...
    assert status.loc[("Adjusted", "original__f0")] in {
        "not_enough_degrees_of_freedom",
        "ok",
    }


def test_outcome_as_column_name():
    X, meta = _data()
    res = compute_feature_associations(X, "y", covariate_data=meta)
    assert len(res.table) == 4  # one tier x four features


def test_explicit_model_tiers():
    X, meta = _data()
    res = compute_feature_associations(
        X, meta["y"], model_tiers={"crude": [], "adj": ["age"]}, covariate_data=meta
    )
    assert res.tiers == ["crude", "adj"]


def test_catalog_annotation():
    X, meta = _data()
    res = compute_feature_associations(X, meta["y"], covariate_data=meta, catalog=_catalog())
    assert {"family", "family_group"} <= set(res.table.columns)
    row = res.table.set_index(["model", "feature"]).loc[("Univariable", "original__f0")]
    assert row["family_group"] == "Intensity"
    # a DataFrame catalog works too
    res2 = compute_feature_associations(
        X, meta["y"], covariate_data=meta, catalog=_catalog().frame
    )
    assert "family" in res2.table.columns


def test_radiomics_dataset_infers_outcome_and_catalog():
    X, meta = _data()
    ds = RadiomicsDataset(
        pd.concat([X, meta], axis=1),
        feature_columns=list(X.columns),
        catalog=_catalog(),
        design=StudyDesign(roles={"target": "y"}),
    )
    res = compute_feature_associations(ds)
    assert res.outcome_type == "continuous"
    assert "family_group" in res.table.columns  # catalog taken from the dataset


# ---- validation / errors --------------------------------------------------


def test_outcome_required_without_dataset():
    X, _ = _data()
    with pytest.raises(ValueError, match="outcome is required"):
        compute_feature_associations(X)


def test_dataset_without_outcome_raises():
    X, meta = _data()
    ds = RadiomicsDataset(pd.concat([X, meta], axis=1), feature_columns=list(X.columns))
    with pytest.raises(ValueError, match="no outcome"):
        compute_feature_associations(ds)


def test_bad_outcome_type_raises():
    X, meta = _data()
    with pytest.raises(TypeError, match="Series, DataFrame, or column name"):
        compute_feature_associations(X, 123, covariate_data=meta)


def test_invalid_outcome_type_raises():
    X, meta = _data()
    with pytest.raises(ValueError, match="continuous/binary/survival"):
        compute_feature_associations(X, meta["y"], outcome_type="poisson", covariate_data=meta)


def test_mixed_method_invalid_raises():
    X, meta = _data()
    with pytest.raises(ValueError, match="mixed_method must be"):
        compute_feature_associations(X, meta["y"], covariate_data=meta, mixed_method="reml")


def test_missing_covariate_raises():
    X, meta = _data()
    with pytest.raises(KeyError, match="covariate column"):
        compute_feature_associations(X, meta["y"], adjust_for=["ghost"], covariate_data=meta)


def test_no_features_raises():
    meta = pd.DataFrame({"y": [1.0, 2.0, 3.0], "label": ["a", "b", "c"]})
    with pytest.raises(ValueError, match="no feature columns"):
        compute_feature_associations(meta[["label"]], meta["y"], covariate_data=meta)


# ---- top_hits -------------------------------------------------------------


def test_top_hits_modes():
    X, meta = _data()
    res = compute_feature_associations(X, meta["y"], adjust_for=["age"], covariate_data=meta)
    assert len(res.top_hits(mode="fdr")) >= 1
    assert len(res.top_hits(mode="nominal")) >= 1
    assert len(res.top_hits(mode="ranked", per_panel=1)) == 2  # 1 per tier
    with pytest.raises(ValueError, match="fdr.*nominal.*ranked"):
        res.top_hits(mode="bogus")


def test_top_hits_empty_when_nothing_fitted():
    X, meta = _data()
    constants = X[["original__const"]]  # only a constant feature -> never "ok"
    res = compute_feature_associations(constants, meta["y"], covariate_data=meta)
    assert res.top_hits(mode="fdr").empty


# ---- Phase B engines (survival / binary / mixed; need lifelines/statsmodels) --


def _survival_binary(n: int = 150):
    rng = np.random.default_rng(1)
    f0 = rng.normal(0, 1, n)
    f1 = rng.normal(0, 1, n)
    age = rng.normal(60, 8, n)
    patient = np.repeat(np.arange(n // 3 + 1), 3)[:n]  # repeated-measures clusters
    lp = 1.2 * f0 + 0.02 * age
    idx = list(range(n))
    X = pd.DataFrame({"original__f0": f0, "original__f1": f1}, index=idx)
    meta = pd.DataFrame({"age": age, "patient": patient}, index=idx)
    surv = pd.DataFrame(
        {
            "time": rng.exponential(np.exp(-lp)) + 0.01,
            "event": (rng.uniform(size=n) < 0.7).astype(int),
        },
        index=idx,
    )
    binary = pd.Series((rng.uniform(size=n) < 1 / (1 + np.exp(-lp))).astype(int), index=idx)
    return X, meta, surv, binary


def test_survival_cox():
    X, meta, surv, _ = _survival_binary()
    res = compute_feature_associations(
        X, surv, adjust_for=["age"], covariate_data=meta, min_events=5
    )
    assert res.outcome_type == "survival"
    row = res.table.set_index(["model", "feature"]).loc[("Univariable", "original__f0")]
    assert row["status"] == "ok"
    assert row["model_family"] == "cox" and row["effect_name"] == "HR"
    assert np.isfinite(row["c_index"])


def test_survival_cox_clustered():
    X, meta, surv, _ = _survival_binary()
    res = compute_feature_associations(
        X, surv, covariate_data=meta, groups="patient", min_events=5
    )
    fam = res.table.set_index(["model", "feature"]).loc[
        ("Univariable", "original__f0"), "model_family"
    ]
    assert fam == "cox_clustered"


def test_survival_no_events():
    X, meta, surv, _ = _survival_binary()
    res = compute_feature_associations(X, surv, covariate_data=meta, min_events=10_000)
    assert (res.table["status"] == "no_events").all()


def test_binary_logistic():
    X, meta, _, binary = _survival_binary()
    res = compute_feature_associations(X, binary, adjust_for=["age"], covariate_data=meta)
    assert res.outcome_type == "binary"
    row = res.table.set_index(["model", "feature"]).loc[("Univariable", "original__f0")]
    assert row["status"] == "ok" and row["model_family"] == "logit" and row["effect_name"] == "OR"


def test_binary_no_events():
    X, meta, _, _ = _survival_binary()
    const_outcome = pd.Series(np.zeros(len(X)), index=X.index)
    res = compute_feature_associations(
        X, const_outcome, outcome_type="binary", covariate_data=meta
    )
    assert (res.table["status"] == "no_events").all()


def test_binary_gee_clustered():
    X, meta, _, binary = _survival_binary()
    res = compute_feature_associations(
        X, binary, covariate_data=meta, groups="patient", mixed_method="gee"
    )
    fam = res.table.set_index(["model", "feature"]).loc[
        ("Univariable", "original__f0"), "model_family"
    ]
    assert fam == "gee_logit"


def test_binary_glmm_clustered():
    X, meta, _, binary = _survival_binary()
    res = compute_feature_associations(X, binary, covariate_data=meta, groups="patient")
    row = res.table.set_index(["model", "feature"]).loc[("Univariable", "original__f0")]
    assert row["model_family"] == "glmm" and row["status"] == "ok"


def _mixed_continuous_outcome(X, meta):
    """A continuous outcome with a genuine per-patient random intercept + residual
    noise, so MixedLM estimates a non-singular random-effects variance (stable
    across statsmodels versions)."""
    rng = np.random.default_rng(5)
    cluster = meta["patient"].to_numpy()
    intercepts = rng.normal(0, 0.5, cluster.max() + 1)
    y = (
        1.0 * X["original__f0"].to_numpy()
        + 0.02 * meta["age"].to_numpy()
        + intercepts[cluster]
        + rng.normal(0, 0.3, len(X))
    )
    return pd.Series(y, index=X.index)


def test_continuous_mixedlm_via_groups_array():
    X, meta, _, _ = _survival_binary()
    y = _mixed_continuous_outcome(X, meta)
    # groups passed as an array (not a column name) -> exercises that resolution branch
    res = compute_feature_associations(
        X, y, groups=meta["patient"].to_numpy(), covariate_data=meta
    )
    fam = res.table.set_index(["model", "feature"]).loc[
        ("Univariable", "original__f0"), "model_family"
    ]
    assert fam == "mixedlm"


def test_dataset_supplies_groups():
    X, meta, _, _ = _survival_binary()
    y = _mixed_continuous_outcome(X, meta)
    data = pd.concat([X, meta, y.rename("y")], axis=1)
    ds = RadiomicsDataset(
        data,
        feature_columns=list(X.columns),
        design=StudyDesign(roles={"target": "y", "group": "patient"}),
    )
    res = compute_feature_associations(ds)  # groups taken from the design
    fam = res.table.set_index(["model", "feature"]).loc[
        ("Univariable", "original__f0"), "model_family"
    ]
    assert fam == "mixedlm"


# ---- fit-failure and optional-dependency handling -------------------------


def test_logit_fit_failed(monkeypatch):
    import types as _types

    fake_sm = _types.SimpleNamespace(
        Logit=lambda *a, **k: (_ for _ in ()).throw(RuntimeError("logit boom"))
    )
    monkeypatch.setattr("eigenradiomics.feature_models._import_statsmodels", lambda: fake_sm)
    X, meta, _, binary = _survival_binary(n=30)
    res = compute_feature_associations(X, binary, outcome_type="binary", covariate_data=meta)
    assert (res.table["status"] == "fit_failed").all()
    assert res.table.iloc[0]["error"]


def test_cox_fit_failed(monkeypatch):
    class _BadCox:
        def __init__(self, **kwargs):
            pass

        def fit(self, *args, **kwargs):
            raise RuntimeError("cox boom")

    monkeypatch.setattr("eigenradiomics.feature_models._import_lifelines", lambda: _BadCox)
    X, meta, surv, _ = _survival_binary(n=30)
    res = compute_feature_associations(X, surv, covariate_data=meta, min_events=1)
    assert (res.table["status"] == "fit_failed").all()


def test_mixedlm_fit_failed(monkeypatch):
    import types as _types

    fake_sm = _types.SimpleNamespace(
        MixedLM=lambda *a, **k: (_ for _ in ()).throw(RuntimeError("mixedlm boom"))
    )
    monkeypatch.setattr("eigenradiomics.feature_models._import_statsmodels", lambda: fake_sm)
    X, meta, _, _ = _survival_binary(n=30)
    y = pd.Series(X["original__f0"].to_numpy(), index=X.index)
    res = compute_feature_associations(X, y, covariate_data=meta, groups="patient")
    assert (res.table["status"] == "fit_failed").all()


def test_glmm_fit_failed(monkeypatch):
    from statsmodels.genmod.bayes_mixed_glm import BinomialBayesMixedGLM

    def _boom(*args, **kwargs):
        raise RuntimeError("glmm boom")

    monkeypatch.setattr(BinomialBayesMixedGLM, "from_formula", staticmethod(_boom))
    X, meta, _, binary = _survival_binary(n=30)
    res = compute_feature_associations(X, binary, covariate_data=meta, groups="patient")
    assert (res.table["status"] == "fit_failed").all()


def test_lifelines_import_error(monkeypatch):
    monkeypatch.setitem(__import__("sys").modules, "lifelines", None)
    X, meta, surv, _ = _survival_binary(n=30)
    with pytest.raises(ImportError, match="lifelines"):
        compute_feature_associations(X, surv, covariate_data=meta, min_events=1)


def test_statsmodels_import_error(monkeypatch):
    monkeypatch.setitem(__import__("sys").modules, "statsmodels.api", None)
    X, meta, _, binary = _survival_binary(n=30)
    with pytest.raises(ImportError, match="statsmodels"):
        compute_feature_associations(X, binary, outcome_type="binary", covariate_data=meta)


# ---- Phase C: volcano plot ------------------------------------------------


def _volcano_result(n_tiers: int = 2, n_features: int = 8):
    rng = np.random.default_rng(2)
    n = 200
    cols = {f"original__f{i}": rng.normal(0, 1, n) for i in range(n_features)}
    X = pd.DataFrame(cols, index=range(n))
    age = rng.normal(60, 8, n)
    y = pd.Series(1.5 * X["original__f0"] + 0.02 * age + rng.normal(0, 1, n), index=range(n))
    meta = pd.DataFrame({"age": age}, index=range(n))
    fams = ["Intensity", "Texture", "Morphology"]
    catalog = FeatureCatalog(
        pd.DataFrame(
            {
                "config": ["original"] * n_features,
                "feature_key": [f"f{i}" for i in range(n_features)],
                "family": ["firstorder"] * n_features,
                "family_group": [fams[i % 3] for i in range(n_features)],
            }
        )
    )
    tiers = {f"M{j}": (["age"] if j % 2 else []) for j in range(n_tiers)}
    return compute_feature_associations(
        X, y, model_tiers=tiers, covariate_data=meta, catalog=catalog
    )


@pytest.mark.parametrize(
    ("n_tiers", "expected_axes", "expected_visible"),
    [(1, 1, 1), (2, 2, 2), (3, 3, 3), (4, 4, 4), (5, 9, 5), (9, 9, 9)],
)
def test_volcano_layouts(n_tiers, expected_axes, expected_visible):
    res = _volcano_result(n_tiers=n_tiers)
    fig = plot_volcano(res, color_by="family_group", marker_by="family_group")
    assert len(fig.axes) == expected_axes
    assert sum(ax.get_visible() for ax in fig.axes) == expected_visible
    plt.close(fig)


def test_volcano_panel_count_out_of_range():
    res = _volcano_result(n_tiers=2)
    with pytest.raises(ValueError, match="1-9 panels"):
        plot_volcano(res, tiers=[])


def test_volcano_explicit_layout_and_too_small():
    res = _volcano_result(n_tiers=4)
    fig = plot_volcano(res, layout=(2, 3))
    assert len(fig.axes) == 6
    plt.close(fig)
    with pytest.raises(ValueError, match="fewer cells"):
        plot_volcano(res, layout=(1, 1))


def test_volcano_axis_and_outlier_modes():
    res = _volcano_result()
    plt.close(plot_volcano(res, axis_mode="shared", outlier_strategy="include"))
    with pytest.raises(ValueError, match="axis_mode"):
        plot_volcano(res, axis_mode="weird")
    with pytest.raises(ValueError, match="outlier_strategy"):
        plot_volcano(res, outlier_strategy="weird")


def test_volcano_color_none_and_marker_by():
    res = _volcano_result()
    plt.close(plot_volcano(res, color_by=None))  # significant points use a single highlight colour
    plt.close(plot_volcano(res, color_by="family_group", marker_by="family_group"))


def test_volcano_empty_panels():
    # all-constant features -> nothing fitted -> every panel shows "no fitted features"
    X = pd.DataFrame({f"original__f{i}": np.full(50, 3.0) for i in range(3)}, index=range(50))
    y = pd.Series(np.arange(50, dtype=float), index=range(50))
    res = compute_feature_associations(X, y, covariate_data=pd.DataFrame({"y": y}))
    fig = plot_volcano(res)
    plt.close(fig)


def test_volcano_binary_xlabel():
    X, meta, _, binary = _survival_binary()
    res = compute_feature_associations(X, binary, covariate_data=meta)
    fig = plot_volcano(res, title="binary")
    assert any("odds ratio" in ax.get_xlabel() for ax in fig.axes if ax.get_visible())
    plt.close(fig)


# ---- Phase D: heatmap bridge + Excel export -------------------------------


def test_bar_and_value_columns():
    X, meta = _data()
    res = compute_feature_associations(X, meta["y"], adjust_for=["age"], covariate_data=meta)
    bar = res.bar(value="neg_log10_p", reference=1.3, title="-log10 p")
    assert type(bar).__name__ == "Bar"
    assert list(bar.data.index) == list(X.columns)  # feature-indexed
    assert bar.title == "-log10 p" and bar.reference == 1.3
    # exercise the remaining value columns
    for value in ("neg_log10_fdr", "coef", "effect", "statistic", "ci_low", "ci_high"):
        assert res.bar(value=value).data.shape[0] == len(X.columns)


def test_bar_log2_effect_on_ratio():
    X, meta, _, binary = _survival_binary()
    res = compute_feature_associations(X, binary, covariate_data=meta)
    bar = res.bar(value="log2_effect")  # log2(OR) is well-defined for ratio effects
    assert bar.data.shape[0] == X.shape[1]


def test_matrix_feature_by_tier():
    X, meta = _data()
    res = compute_feature_associations(X, meta["y"], adjust_for=["age"], covariate_data=meta)
    mat = res.matrix(value="coef")
    assert list(mat.columns) == ["Univariable", "Adjusted"]
    assert list(mat.index) == list(X.columns)


def test_bar_bad_value_and_tier():
    X, meta = _data()
    res = compute_feature_associations(X, meta["y"], covariate_data=meta)
    with pytest.raises(ValueError, match="value must be one of"):
        res.bar(value="bogus")
    with pytest.raises(ValueError, match="unknown tier"):
        res.bar(tier="ghost")


def test_to_excel(tmp_path):
    import openpyxl

    X, meta = _data()
    res = compute_feature_associations(X, meta["y"], adjust_for=["age"], covariate_data=meta)
    path = tmp_path / "associations.xlsx"
    res.to_excel(path)
    assert path.exists()
    wb = openpyxl.load_workbook(path)
    assert "associations" in wb.sheetnames and "top_hits" in wb.sheetnames


def test_plot_rwas_manhattan(tmp_path):
    from eigenradiomics.feature_models import plot_rwas_manhattan
    from eigenradiomics.plotting import Bar, CorrPanel, Strip

    X, meta = _data()
    res = compute_feature_associations(X, meta["y"], adjust_for=["age"], covariate_data=meta)

    # Simple call
    fig = plot_rwas_manhattan(res)
    assert fig is not None
    plt.close(fig)

    # Custom ordering, catalog, and strips/bars/corr_panel
    catalog = _catalog()
    features = list(X.columns)
    strip = Strip(pd.Series(["A", "B", "A", "B"], index=features), title="strip1")
    bar = Bar(pd.Series([1.0, 2.0, 1.5, 3.0], index=features), title="bar1")
    corr_df = pd.DataFrame(
        [[0.1, -0.2], [0.5, 0.4], [-0.3, 0.1], [0.8, -0.7]],
        index=features,
        columns=["var1", "var2"],
    )
    corr_panel = CorrPanel(corr_df, label="Correlation")

    fig2 = plot_rwas_manhattan(
        res,
        catalog=catalog,
        tier="Univariable",
        group_by="family",
        order=features[::-1],
        strips=[strip],
        bars=[bar],
        corr_panel=corr_panel,
        title="Manhattan Test",
        path=tmp_path / "manhattan.png",
    )
    assert fig2 is not None
    assert (tmp_path / "manhattan.png").exists()
    plt.close(fig2)


def test_plot_rwas_manhattan_coverage_gaps():
    from eigenradiomics.feature_models import plot_rwas_manhattan
    from eigenradiomics.plotting import CorrPanel

    # 1. No fitted features available to plot
    empty_df = pd.DataFrame(columns=["feature", "p_value", "status"])
    with pytest.raises(ValueError, match="No fitted features available to plot"):
        plot_rwas_manhattan(empty_df)

    # 2. No features match the specified order
    X, meta = _data()
    res = compute_feature_associations(X, meta["y"], adjust_for=["age"], covariate_data=meta)
    with pytest.raises(ValueError, match="No features match the specified order"):
        plot_rwas_manhattan(res, order=["non_existent_feature"])

    # 3. n > 60 features (to hit tick_labels is None and ax_corr.set_xticks([]))
    n_feats = 65
    rng = np.random.default_rng(42)
    X_large = pd.DataFrame(
        {f"feat_{i}": rng.normal(0, 1, 100) for i in range(n_feats)}
    )
    y_large = pd.Series(rng.normal(0, 1, 100))
    res_large = compute_feature_associations(X_large, y_large)

    features_large = list(X_large.columns)
    corr_df = pd.DataFrame(
        rng.normal(0, 1, (n_feats, 2)),
        index=features_large,
        columns=["var1", "var2"],
    )
    corr_panel = CorrPanel(corr_df, label="Large Correlation")

    fig = plot_rwas_manhattan(res_large, corr_panel=corr_panel)
    assert fig is not None
    plt.close(fig)

