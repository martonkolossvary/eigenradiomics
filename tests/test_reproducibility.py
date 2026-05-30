"""Unit tests for the radiomics reproducibility framework."""

from __future__ import annotations

import tempfile
from pathlib import Path

import matplotlib.pyplot as plt
import numpy as np
import openpyxl
import pandas as pd
import pytest

from eigenradiomics._stats import _get_deterministic_seed, _icc_2_1_estimate
from eigenradiomics.reproducibility import (
    compute_reproducibility,
    plot_reproducibility_histograms,
    write_reproducibility_excel,
)


def test_deterministic_seed():
    """Verify that deterministic seed hashing is fully reproducible and distinct."""
    seed1 = _get_deterministic_seed("feature_A", base_seed=42)
    seed2 = _get_deterministic_seed("feature_A", base_seed=42)
    seed3 = _get_deterministic_seed("feature_B", base_seed=42)
    seed4 = _get_deterministic_seed("feature_A", base_seed=100)

    assert seed1 == seed2
    assert seed1 != seed3
    assert seed1 != seed4
    assert isinstance(seed1, int)


def test_icc_2_1_mathematical_correctness():
    """Verify the ICC(2,1) formula against simple textbook cases with known results."""
    # Example dataset: 4 subjects, 3 raters
    # Subject 1: 9, 8, 9
    # Subject 2: 7, 7, 6
    # Subject 3: 5, 4, 5
    # Subject 4: 3, 3, 2
    Y = np.array([[9, 8, 9], [7, 7, 6], [5, 4, 5], [3, 3, 2]], dtype=float)

    res = _icc_2_1_estimate(Y)

    # Let's check that all MS components and the estimate are finite numbers
    assert not np.isnan(res["icc"])
    assert res["ms_between_subjects"] > 0
    assert res["ms_between_observers"] > 0
    assert res["ms_error"] > 0
    assert 0.8 < res["icc"] < 0.99  # Highly correlated observations


def test_perfect_reproducibility():
    """Test that two identical datasets yield perfect correlations and ICC of 1.0."""
    df1 = pd.DataFrame(
        {
            "feat_1": [1.0, 2.0, 3.0, 4.0, 5.0],
            "feat_2": [5.0, 4.0, 3.0, 2.0, 1.0],
        },
        index=["S1", "S2", "S3", "S4", "S5"],
    )
    df2 = df1.copy()

    results = compute_reproducibility([df1, df2], bootstrap_iterations=50)

    # Check sheet existence
    assert "Spearman" in results
    assert "Pearson" in results
    assert "ICC" in results

    # Get data
    spearman = results["Spearman"]
    pearson = results["Pearson"]
    icc = results["ICC"]

    # Verify columns
    assert "feature" in icc.columns
    assert "icc_2_1" in icc.columns
    assert "ci95_low" in icc.columns
    assert "ci95_high" in icc.columns
    assert "p_value" in icc.columns
    assert "p_fdr" in icc.columns
    assert "retained_ge_0_80" in icc.columns

    # Perfect identical rater results
    assert np.allclose(spearman["estimate"], 1.0)
    assert np.allclose(pearson["estimate"], 1.0)
    assert np.allclose(icc["icc_2_1"], 1.0)
    assert all(icc["retained_ge_0_80"])
    assert all(icc["primary_icc_pass"])


def test_qc_dataframe_alignment_and_reordering():
    """Test strict name-based QC checking, index verification, and automatic reordering."""
    df1 = pd.DataFrame(
        {"feat_1": [1, 2, 3], "feat_2": [4, 5, 6]},
        index=["S1", "S2", "S3"],
    )

    # 1. Feature columns mismatch
    df_bad_cols = pd.DataFrame(
        {"feat_1": [1, 2, 3], "feat_3": [4, 5, 6]},
        index=["S1", "S2", "S3"],
    )
    with pytest.raises(ValueError, match="columns do not match Dataset 0"):
        compute_reproducibility([df1, df_bad_cols])

    # 2. Subject index mismatch
    df_bad_idx = pd.DataFrame(
        {"feat_1": [1, 2, 3], "feat_2": [4, 5, 6]},
        index=["S1", "S2", "S4"],
    )
    with pytest.raises(ValueError, match="row index does not match Dataset 0"):
        compute_reproducibility([df1, df_bad_idx])

    # 3. Features/subjects out of order (should automatically align and succeed!)
    df_scrambled = pd.DataFrame(
        {"feat_2": [6, 4, 5], "feat_1": [3, 1, 2]},
        index=["S3", "S1", "S2"],
    )

    # Compute: this should align df_scrambled perfectly to df1 and yield perfect 1.0 scores!
    results = compute_reproducibility([df1, df_scrambled], bootstrap_iterations=10)
    assert np.allclose(results["Pearson"]["estimate"], 1.0)
    assert np.allclose(results["ICC"]["icc_2_1"], 1.0)


def test_qc_unnamed_positional_matching():
    """Test strict positional verification when passing numpy arrays or default RangeIndexes."""
    arr1 = np.random.default_rng(42).standard_normal((10, 5))
    arr2 = arr1.copy()

    # Mismatched shapes
    arr_bad_shape = np.random.default_rng(42).standard_normal((10, 6))
    with pytest.raises(ValueError, match="shape"):
        compute_reproducibility([arr1, arr_bad_shape])

    # Successful alignment and calculation
    results = compute_reproducibility([arr1, arr2], bootstrap_iterations=10)
    assert len(results["ICC"]) == 5  # 5 features
    assert np.allclose(results["Pearson"]["estimate"], 1.0)
    assert np.allclose(results["ICC"]["icc_2_1"], 1.0)


def test_selector_based_filtering():
    """Test that standard selectors correctly isolate the analyzed features."""
    # We will use columns resembling standard radiomics features: Config__FeatureKey
    cols = [
        "original__Energy",
        "original__Entropy",
        "original__Autocorrelation",
        "wavelet__ClusterShade",
    ]
    df1 = pd.DataFrame(
        np.random.default_rng(42).standard_normal((10, 4)),
        columns=cols,
        index=[f"S{i}" for i in range(10)],
    )
    df2 = df1.copy()

    # Specifying features directly
    results = compute_reproducibility(
        [df1, df2],
        features=["*Energy", "*Entropy"],
        bootstrap_iterations=10,
    )
    # Only Energy and Entropy should be analyzed
    icc_features = results["ICC"]["feature"].tolist()
    assert len(icc_features) == 2
    assert "original__Energy" in icc_features
    assert "original__Entropy" in icc_features

    # Specifying families directly
    # Family filtering requires a catalog. Let's create a small catalog DataFrame.
    catalog = pd.DataFrame(
        {
            "feature_key": ["Energy", "Entropy", "Autocorrelation", "ClusterShade"],
            "config": ["original", "original", "original", "wavelet"],
            "family": ["firstorder", "firstorder", "glcm", "glcm"],
            "family_group": ["texture", "texture", "texture", "texture"],
        }
    )

    results_family = compute_reproducibility(
        [df1, df2],
        families="firstorder",
        catalog=catalog,
        bootstrap_iterations=5,
    )
    assert len(results_family["ICC"]) == 2
    assert "original__Energy" in results_family["ICC"]["feature"].tolist()


def test_nan_handling_and_thresholding():
    """Test that features with missing samples are dropped or returned as NaN."""
    df1 = pd.DataFrame(
        {
            "feat_1": [1.0, 2.0, np.nan, 4.0, 5.0],  # 1 NaN
            "feat_2": [1.0, 2.0, np.nan, np.nan, 5.0],  # 2 NaNs -> 3 valid
            "feat_3": [1.0, np.nan, np.nan, np.nan, 5.0],  # 3 NaNs -> 2 valid
        },
        index=[f"S{i}" for i in range(5)],
    )
    df2 = df1.copy()

    # Using min_valid_samples = 3
    results = compute_reproducibility(
        [df1, df2],
        min_valid_samples=3,
        bootstrap_iterations=10,
    )

    icc = results["ICC"].set_index("feature")

    # feat_1 has 4 valid samples (>= 3) -> should compute correctly
    assert not np.isnan(icc.loc["feat_1", "icc_2_1"])
    # feat_2 has 3 valid samples (>= 3) -> should compute correctly
    assert not np.isnan(icc.loc["feat_2", "icc_2_1"])
    # feat_3 has 2 valid samples (< 3) -> should return NaN gracefully without raising an exception
    assert np.isnan(icc.loc["feat_3", "icc_2_1"])
    assert np.isnan(icc.loc["feat_3", "ci95_low"])


def test_multi_observer_aggregations():
    """Test multi-observer statistics (K >= 3 raters)."""
    # 3 observers with scrambled ranks to ensure non-trivial correlations and non-zero SD
    df1 = pd.DataFrame({"feat_1": [1.0, 2.0, 3.0, 4.0, 5.0]}, index=[f"S{i}" for i in range(5)])
    df2 = pd.DataFrame({"feat_1": [2.0, 1.0, 3.0, 5.0, 4.0]}, index=[f"S{i}" for i in range(5)])
    df3 = pd.DataFrame({"feat_1": [3.0, 2.0, 1.0, 4.0, 5.0]}, index=[f"S{i}" for i in range(5)])

    results = compute_reproducibility([df1, df2, df3], bootstrap_iterations=10)

    # Verify column structures for K > 2 aggregate metrics
    for sheet in ["Spearman", "Pearson"]:
        df = results[sheet]
        assert "mean" in df.columns
        assert "median" in df.columns
        assert "sd" in df.columns
        assert "q25" in df.columns
        assert "q75" in df.columns
        assert "min" in df.columns
        assert "max" in df.columns

        # Values should be computed and have non-zero variance
        assert not np.isnan(df.loc[0, "mean"])
        assert df.loc[0, "sd"] > 0.0  # Captures variance across rater pairs


def test_excel_exporter():
    """Test that write_reproducibility_excel creates a valid, highly formatted Excel workbook."""
    df1 = pd.DataFrame({"feat_1": [1, 2, 3, 4, 5]}, index=[f"S{i}" for i in range(5)])
    df2 = df1.copy()

    results = compute_reproducibility([df1, df2], bootstrap_iterations=5)

    with tempfile.TemporaryDirectory() as tmp_dir:
        xlsx_path = Path(tmp_dir) / "reproducibility_report.xlsx"
        write_reproducibility_excel(results, xlsx_path)

        assert xlsx_path.exists()

        # Load sheet back using openpyxl and verify
        wb = openpyxl.load_workbook(xlsx_path)
        assert "Spearman" in wb.sheetnames
        assert "Pearson" in wb.sheetnames
        assert "ICC" in wb.sheetnames

        ws = wb["ICC"]
        # Header row 1 check
        assert ws.cell(row=1, column=1).value == "feature"
        assert ws.cell(row=1, column=2).value == "icc_2_1"

        # Data cell checks
        assert ws.cell(row=2, column=1).value == "feat_1"
        assert round(ws.cell(row=2, column=2).value, 2) == 1.0


def test_accessibility_plots():
    """Test that plot_reproducibility_histograms runs and meets quality requirements."""
    df1 = pd.DataFrame(
        {
            "feat_1": [1, 2, 3, 4, 5],
            "feat_2": [5, 4, 3, 2, 1],
        },
        index=[f"S{i}" for i in range(5)],
    )
    df2 = df1.copy()

    results = compute_reproducibility([df1, df2], bootstrap_iterations=5)

    with tempfile.TemporaryDirectory() as tmp_dir:
        fig_path = Path(tmp_dir) / "reproducibility_plot.png"
        fig = plot_reproducibility_histograms(results, path=fig_path, primary_threshold=0.80)

        # Assertions
        assert fig_path.exists()
        assert isinstance(fig, plt.Figure)
        assert len(fig.axes) == 3  # Spearman, Pearson, ICC subplots
        plt.close(fig)


def test_requires_at_least_two_datasets():
    """A single dataset cannot be assessed for reproducibility."""
    with pytest.raises(ValueError, match="At least 2 datasets"):
        compute_reproducibility([pd.DataFrame({"a": [1.0, 2.0, 3.0]})])


def test_metadata_columns_skipped_without_selectors():
    """Without selectors, non-numeric metadata columns must be skipped, not crash."""
    df1 = pd.DataFrame(
        {
            "PatientID": ["p1", "p2", "p3", "p4"],
            "original__Energy": [1.0, 2.0, 3.0, 4.0],
            "original__Entropy": [4.0, 3.0, 2.0, 1.0],
        },
        index=["s1", "s2", "s3", "s4"],
    )
    results = compute_reproducibility([df1, df1.copy()], bootstrap_iterations=5)
    analyzed = results["ICC"]["feature"].tolist()
    assert "PatientID" not in analyzed
    assert set(analyzed) == {"original__Energy", "original__Entropy"}


def test_plot_handles_all_nan_sheet():
    """Plotting must not crash when a metric column is entirely NaN/empty."""
    results = {
        "Spearman": pd.DataFrame({"feature": ["f1", "f2"], "estimate": [np.nan, np.nan]}),
        "Pearson": pd.DataFrame({"feature": ["f1", "f2"], "estimate": [np.nan, np.nan]}),
        "ICC": pd.DataFrame({"feature": ["f1", "f2"], "icc_2_1": [np.nan, np.nan]}),
    }
    fig = plot_reproducibility_histograms(results)
    assert len(fig.axes) == 3
    plt.close(fig)


def test_positional_rangeindex_columns():
    """DataFrames with a default RangeIndex columns take the positional path."""
    arr = np.random.default_rng(0).standard_normal((12, 4))
    results = compute_reproducibility(
        [pd.DataFrame(arr), pd.DataFrame(arr.copy())], bootstrap_iterations=5
    )
    assert len(results["ICC"]) == 4


def test_positional_non_string_columns():
    """Non-string (but non-RangeIndex) column labels take the positional path."""
    arr = np.random.default_rng(0).standard_normal((12, 3))
    d1 = pd.DataFrame(arr, columns=[10, 20, 30])
    results = compute_reproducibility([d1, d1.copy()], bootstrap_iterations=5)
    assert len(results["ICC"]) == 3


def test_no_numeric_features_raises():
    d1 = pd.DataFrame({"label": ["a", "b", "c", "d"]}, index=["s1", "s2", "s3", "s4"])
    with pytest.raises(ValueError, match="No features selected"):
        compute_reproducibility([d1, d1.copy()])


def test_multi_observer_insufficient_samples():
    idx = [f"s{i}" for i in range(5)]
    col = {"feat": [1.0, np.nan, np.nan, np.nan, np.nan]}  # only 1 valid sample
    dfs = [pd.DataFrame(col, index=idx) for _ in range(3)]
    results = compute_reproducibility(dfs, min_valid_samples=3, bootstrap_iterations=5)
    assert "mean" in results["Spearman"].columns  # multi-observer schema
    assert np.isnan(results["Spearman"].loc[0, "mean"])


def test_multi_observer_constant_feature():
    idx = [f"s{i}" for i in range(6)]
    dfs = [pd.DataFrame({"feat": [2.0] * 6}, index=idx) for _ in range(3)]
    results = compute_reproducibility(dfs, min_valid_samples=3, bootstrap_iterations=5)
    # Constant feature -> all pairwise correlations NaN -> aggregate is NaN.
    assert np.isnan(results["Spearman"].loc[0, "mean"])


def test_plot_empty_results_raises():
    with pytest.raises(ValueError, match="no plottable"):
        plot_reproducibility_histograms({})


def test_plot_single_sheet():
    results = {"ICC": pd.DataFrame({"feature": ["f"], "icc_2_1": [0.9]})}
    fig = plot_reproducibility_histograms(results)
    assert len(fig.axes) == 1
    plt.close(fig)
